import argparse
from datetime import datetime
from openpyxl import load_workbook
import os
import subprocess

from chemicalio import map_species_to_indices, map_single_species
from reactions import *

def createPajekConfiguration (chemicalSpecies, reactions):

    #* path directory definition
    directory_name = "../out"

    if not os.path.exists(directory_name):
        try:
            os.makedirs(directory_name)
        except subprocess.CalledProcessError as e:
            print(f"Error in creating the directory: {e}")

    currentData = datetime.now().strftime("%d.%m")
    directory_name = f"../out/pajek {currentData}"
    
    if not os.path.exists(directory_name):
        try:
            os.makedirs(directory_name)
        except subprocess.CalledProcessError as e:
            print(f"Error in creating second directory: {e}")

    currentTime = datetime.now().strftime("%H.%M.%S")
    fileName = f"../out/{directory_name}/graphSim {currentTime} .paj"

    with open(fileName, 'w') as file:
            
        i = 0
        file.write ("Network Bipartite directed Erdos-Renyi random network\n")
        file.write (f"*Vertices {len(chemicalSpecies)+len(reactions)}\n")

        for item in chemicalSpecies:
            i+=1
            file.write(f"\t{i} \"{item}\"\t ellipse ic Yellow\n")
            
        for k , item in enumerate(reactions):
            i+=1
            file.write(f"\t{i} \"R{k+1}\"\t box ic Red\n")

        '''ll=1 
        iR = 1
        ch = len(chemicalSpecies)
        file.write ("*Arcs\n")
        for item in reactions:
            for iStart in item["in"]:
                for iOut in item["out"]: 
                    file.write (f'\t{iStart} {iR+ch} {ll}\n')
                    file.write (f'\t{iR+ch} {iOut} {ll}\n')
            iR+=1'''

        ll=1 
        iR = 1
        ch = len(chemicalSpecies)
        file.write ("*Arcs\n")
        for item in reactions:
            cataList =item["catalyst"]

            if item ["k"] != "None":
                ll = item ["k"]
            else:
                ll = "1"

            for cat, iCat in cataList:

                index = map_single_species (cat, chemicalSpecies)
                index += 1

                if iCat == 0: 
                    file.write (f'\t{index} {iR+ch} {-ll} c Blue\n')
                
                if iCat > 0: 
                    file.write (f'\t{index} {iR+ch} {-ll} c Black\n')

                if iCat < 0: 
                    file.write (f'\t{index} {iR+ch} {-ll} c Red\n')

            for iStart in item["in"]:
                for iOut in item["out"]: 
                    file.write (f'\t{iStart} {iR+ch} {ll}\n')
                    file.write (f'\t{iR+ch} {iOut} {ll}\n')
            iR+=1

        file.write ("*Partition Partition into two subsets in N1\n")
        file.write (f"*Vertices {len(chemicalSpecies)+len(reactions)}\n")
        for i in enumerate (chemicalSpecies):
            file.write ("1\n")
        for i in enumerate (reactions):
            file.write ("2\n")
                      
def increase_indices_by_one(reactions_list):
    for reaction in reactions_list:
        if reaction["in"] is not None:
            reaction["in"] = [idx + 1 if idx is not None else None for idx in reaction["in"]]
        if reaction["out"] is not None:
            reaction["out"] = [idx + 1 if idx is not None else None for idx in reaction["out"]]

def reactionsCleaning (inReactions): 

    reactions = inReactions

    for reaction in reactions: 
        cataList = reaction ["catalyst"]
        for catalyst, indexCata in cataList:
            
            if indexCata == 0: # species only catalyzes reaction
                reaction["in"].remove(catalyst)
                reaction["out"].remove(catalyst)
            
            if indexCata > 0: # species catalyzes and is produced in reaction
                reaction["in"].remove(catalyst)
                reaction["out"].remove(catalyst)
            
            if indexCata < 0: # species catalyzes and reacts in reaction
                reaction["in"].remove(catalyst)
                reaction["out"].remove(catalyst)

    return reactions

def loadInfoFromSim (fileName, verbose):

    workbook = load_workbook (fileName)
    
    #* Fluxes reading
    sheet = workbook['Environment']
    fluxParameter = int (sheet.cell(row=7, column=2).value)
    nIteratesParameter = int (sheet.cell(row=10, column=2).value)

    if verbose: 
        print ("flux recognized: ", fluxParameter)
        print ("iterates recognized: ", nIteratesParameter)
    
    sheetFlux = None
    if fluxParameter > 0: 
        generationNumber = int(input ("Type the generation number to view the fluxes> "))
        checkProtoSim(7, [[generationNumber], nIteratesParameter, -1])
        sheetFlux = workbook['Flux']
        generationNumber+=1
    
    #* Chemical Species Reading
    sheet = workbook['Chemical Species']
    chemicalSpecies = [item.value for column in sheet.iter_cols(min_col=1, max_col=1) for item in column][1:]
   
    if verbose:
        print (f"Chemical species [{len(chemicalSpecies)}] : ", end="")
        for item in chemicalSpecies: 
            print (item, end= " ")
        print ("\n")

    #* Reactions Reading
    sheet = workbook['Reactions']
    lines = []

    for cell in sheet.iter_rows(values_only=True, min_row=2):
        
        if cell[4] == 'Diffusion' or cell[4] == 'in-CSTR' or cell[4] == 'CSTR-out':
            continue
        
        line = f"{cell[0]} > {cell[2]}"
        lines.append(line)
    
    if verbose: 
        print ("reaction lines imported: ", lines)

    reactions = []
    for i, reaction in enumerate(lines, start=1):
        
        type = identifyType (reaction, verbose)
        catalyst = identifyCatalysts (reaction, verbose)
        
        reagents, products = reaction.split('>')
        reagents = [reagent.strip() for reagent in reagents.split('+')]
        products = [product.strip() for product in products.split('+')]

        iReaction = i

        fluxReaction = "None"
        if fluxParameter > 0:
            if iReaction <= fluxParameter: 
                fluxReaction = sheetFlux.cell(row=generationNumber+1, column=iReaction+2).value

        reaction_data = {
                        "in": reagents,
                        "out": products,
                        "type": type,
                        "k": fluxReaction,
                        "catalyst" : catalyst
                    }
        
        if verbose: 
            print ("]reaction index: ", iReaction)
            print ("reactants: ", reagents, "|products: ", products, "|type: ", type.value, "|flux: ", fluxReaction, "|catalyst: ", catalyst, "\n")

        reactions.append(reaction_data)

    cReactions = reactionsCleaning (reactions)

    index = map_species_to_indices (cReactions, chemicalSpecies)
    increase_indices_by_one(index)

    for rw, rW in zip(cReactions, index):
        if "catalyst" in rw:
            rW["catalyst"] = rw["catalyst"]

    '''for item in index: 
        print ("start: ", item["in"], "dest: ", item["out"])
    
    for item in index: 
        for iStart in item["in"]:
            for iOut in item ["out"]: 
                print ("start: ", iStart, "dest: ", iOut)'''

    return chemicalSpecies, index

if __name__ == "__main__":

    parser = argparse.ArgumentParser(description="pajeko v1 - 284660")
    parser.add_argument("-v", "--verbose", action="store_true", help="additional prints")
    parser.add_argument("-f", "--file", action="store_true", help="specify input file")

    args = parser.parse_args()
    
    fileName = "../input/pajeko/"
    
    if args.file: 
        fileUser = input ("Type the name of .xlsx file> ")
        fileName = fileName + fileUser
    else: 
        fileName = fileName + "sim.xlsx"

    chemicalSpecies, reactions = loadInfoFromSim (fileName, args.verbose)
    createPajekConfiguration (chemicalSpecies, reactions)
